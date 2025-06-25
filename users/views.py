from django.shortcuts import render, redirect,get_object_or_404, redirect
from django.contrib.auth import login,logout,authenticate
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .forms import CustomUserCreationForm
from. forms import ProfileForm,Level1ProfileForm,Level2ProfileForm,Level3ProfileForm,CompanyProfileForm,MessageForm
from . models import Level1Profile,Level2Profile,Level3Profile, CompanyProfile,Message,Profile,Address,Skill,Category
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
#location fields
import json
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.contrib.auth.models import User
from users.models import Level1Profile, Level2Profile, Level3Profile, CompanyProfile, Skill, Address
from django.db import models
import csv
import io
import openpyxl
from django.http import HttpResponse
from django.utils.timezone import is_aware
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from django.templatetags.static import static
from datetime import datetime

def index_view(request):
    return render(request, 'index.html')
#_____________________________User authentication _______________________________________

#Login a user in

def login_view(request):
    if request.method == 'POST':
        login_form = AuthenticationForm(request=request,data=request.POST)
        if login_form.is_valid():
            username = login_form.cleaned_data.get('username')
            password = login_form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request,user)
                messages.success(request,f'You are logged in as {username}.')
                return redirect('home')
            else:
                messages.error(request,'Something went wrong!')

        else:
            messages.error(request,f'An error occured!')
    else:
        login_form = AuthenticationForm()

    return render(request,'users/login.html',{'login_form':login_form})

# Logout a user

def logout_view(request):
    logout(request)
    return redirect('register')

# Registering a User

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user=form.save()
            login(request, user) 
             # Get the profile type from the form submission
            profile_type = request.POST.get('profile_type')

            # Redirect based on the chosen profile type
            if profile_type == 'level1':
                return redirect('level1_dashboard')  # Update with actual URL name
            elif profile_type == 'level2':
                return redirect('level2_dashboard')
            elif profile_type == 'level3':
                return redirect('level3_dashboard')
            elif profile_type == 'company':
                return redirect('company_dashboard')
            else:
                return redirect('landing-page')  # Default fallback  

                
    else:
        form = CustomUserCreationForm()

    return render(request, 'users/register.html', {
        'form': form,
    })



# ____________________________________Creating profile levels__________________________________________

# Level1_Profile

def level1_dashboard(request):
    with open(settings.BASE_DIR / 'static' / 'rwandaState.json') as f:
        address_data = json.load(f)

    if request.method == 'POST':
        form = Level1ProfileForm(request.POST, request.FILES)
        if form.is_valid():
            # ✅ 1. Read selected address values
            province = request.POST.get('province')
            district = request.POST.get('district')
            sector = request.POST.get('sector')
            cell = request.POST.get('cell')
            village = request.POST.get('village')

            # ✅ 2. Create or get Address instance
            address, created = Address.objects.get_or_create(
                province=province,
                district=district,
                sector=sector,
                cell=cell,
                village=village
            )

            # ✅ 3. Save profile with that address
            level1_profile = form.save(commit=False)
            level1_profile.user = request.user
            level1_profile.address = address  # <- this is crucial
            level1_profile.save()
            form.save_m2m()  # for skills

            return redirect('index')  # or wherever
    else:
        form = Level1ProfileForm()

    return render(request, 'users/level1.html', {
        'form': form,
        'address_json': json.dumps(address_data)
    })



# Level2

def level2_dashboard(request):
    with open(settings.BASE_DIR / 'static' / 'rwandaState.json') as f:
        address_data = json.load(f)

    if request.method == 'POST':
        form = Level2ProfileForm(request.POST, request.FILES)
        if form.is_valid():
            # ✅ 1. Read selected address values
            province = request.POST.get('province')
            district = request.POST.get('district')
            sector = request.POST.get('sector')
            cell = request.POST.get('cell')
            village = request.POST.get('village')

            # ✅ 2. Create or get Address instance
            address, created = Address.objects.get_or_create(
                province=province,
                district=district,
                sector=sector,
                cell=cell,
                village=village
            )

            # ✅ 3. Save profile with that address
            level2_profile = form.save(commit=False)
            level2_profile.user = request.user
            level2_profile.address = address  # <- this is crucial
            level2_profile.save()
            form.save_m2m()  # for skills

            return redirect('home')  # or wherever
    else:
        form = Level2ProfileForm()

    return render(request, 'users/level2.html', {
        'form': form,
        'address_json': json.dumps(address_data)
    })

# def level2_dashboard(request):
#     if request.method == 'POST':
#         form = Level2ProfileForm(request.POST,request.FILES)
#         if form.is_valid():
#             level2_profile = form.save(commit=False)
#             level2_profile.user= request.user
#             level2_profile.save()
#             return redirect('index')
#     else:
#         form = Level2ProfileForm()  

#     return render(request, 'users/level2.html', {'form': form})

# Level3
def level3_dashboard(request):
    with open(settings.BASE_DIR / 'static' / 'rwandaState.json') as f:
        address_data = json.load(f)

    if request.method == 'POST':
        form = Level3ProfileForm(request.POST, request.FILES)
        if form.is_valid():
            # ✅ 1. Read selected address values
            province = request.POST.get('province')
            district = request.POST.get('district')
            sector = request.POST.get('sector')
            cell = request.POST.get('cell')
            village = request.POST.get('village')

            # ✅ 2. Create or get Address instance
            address, created = Address.objects.get_or_create(
                province=province,
                district=district,
                sector=sector,
                cell=cell,
                village=village
            )

            # ✅ 3. Save profile with that address
            level3_profile = form.save(commit=False)
            level3_profile.user = request.user
            level3_profile.address = address  # <- this is crucial
            level3_profile.save()
            form.save_m2m()  # for skills

            return redirect('home')  # or wherever
    else:
        form = Level3ProfileForm()

    return render(request, 'users/level3.html', {
        'form': form,
        'address_json': json.dumps(address_data)
    })

# def level3_dashboard(request):
#     if request.method == 'POST':
#         form = Level3ProfileForm(request.POST,request.FILES)
#         if form.is_valid():
#             level3_profile = form.save(commit=False)
#             level3_profile.user = request.user
#             level3_profile.save()
#             return redirect('index')

#     else:
#          form = Level3ProfileForm()
  
   
#     return render(request, 'users/level3.html',{'form':form})


# Company
def company_dashboard(request):
    with open(settings.BASE_DIR / 'static' / 'rwandaState.json') as f:
        address_data = json.load(f)

    if request.method == 'POST':
        form = CompanyProfileForm(request.POST, request.FILES)
        if form.is_valid():
            # Step 1: Extract address values
            province = request.POST.get('province')
            district = request.POST.get('district')
            sector = request.POST.get('sector')
            cell = request.POST.get('cell')
            village = request.POST.get('village')

            # Step 2: Get or create the address
            address, created = Address.objects.get_or_create(
                province=province,
                district=district,
                sector=sector,
                cell=cell,
                village=village
            )

            # Step 3: Save profile with address
            company_profile = form.save(commit=False)
            company_profile.user = request.user
            company_profile.address = address
            company_profile.save()

            return redirect('index')  # Or wherever you'd like
    else:
        form = CompanyProfileForm()

    return render(request, 'users/company.html', {
        'form': form,
        'address_json': json.dumps(address_data)
    })


# def company_dashboard(request):
#     if request.method == 'POST':
#         form = CompanyProfileForm(request.POST,request.FILES)
#         if form.is_valid():
#             company_profile = form.save(commit=False)
#             company_profile.user = request.user
#             company_profile.save()
#             return redirect('index')
#     else:

#         form =  CompanyProfileForm()
    
    
#     return render(request, 'users/company.html',{'form':form})


def user_profile_view(request):
    return render(request,'profile_user.html')

def company_profile_view(request):
    return render(request,'profile_company.html')

def profile2_view(request):
    return render(request,'profile_2.html')

def profile1_view(request):

    return render(request,'profile_1.html')

def profile3_view(request):
    return render(request,'profile_3.html')

def home_view(request):
    return render(request,'home.html')

def appoitment_view(request):
    return render(request,'appoitment.html')
def market_place_view(request):
    return render(request,'market_place.html')
@staff_member_required
def superuser_dashboard(request):
    # Key metrics
    total_users = User.objects.count()
    total_level1 = Level1Profile.objects.count()
    total_level2 = Level2Profile.objects.count()
    total_level3 = Level3Profile.objects.count()
    total_companies = CompanyProfile.objects.count()
    total_profiles = total_level1 + total_level2 + total_level3

    # Example: Revenue (replace with your real logic)
    total_revenue = 8450000  # Example static value

    # Top skill
    top_skill = (
        Skill.objects.annotate(num_profiles=models.Count('level1_profiles') + models.Count('level2_profiles') + models.Count('level3_profiles'))
        .order_by('-num_profiles')
        .first()
    )
    top_skill_name = top_skill.name if top_skill else "N/A"
    top_skill_percent = 32  # Example static value

    # Top location
    top_location = (
        Address.objects.values('district')
        .annotate(num=models.Count('id'))
        .order_by('-num')
        .first()
    )
    top_location_name = top_location['district'] if top_location else "N/A"
    top_location_percent = 45  # Example static value

    # Chart data (replace with real queries as needed)
    profile_levels_chart = {
        'labels': ['Level 1', 'Level 2', 'Level 3', 'Companies'],
        'data': [total_level1, total_level2, total_level3, total_companies],
    }
    location_chart = {
        'labels': ['Kigali', 'Huye', 'Musanze', 'Rubavu', 'Rwamagana'],
        'data': [120, 50, 30, 25, 15],  # Replace with real counts
    }
    top_skills_chart = {
        'labels': ['Plumbing', 'Catering', 'Mechanic', 'Electrician', 'Design'],
        'data': [60, 45, 30, 25, 20],  # Replace with real counts
    }
    age_distribution_chart = {
        'labels': ['18-24', '25-34', '35-44', '45+'],
        'data': [70, 90, 40, 20],  # Replace with real counts
    }

    context = {
        'total_users': total_users,
        'total_revenue': total_revenue,
        'top_skill': top_skill_name,
        'top_skill_percent': top_skill_percent,
        'top_location': top_location_name,
        'top_location_percent': top_location_percent,
        'profile_levels_chart': profile_levels_chart,
        'location_chart': location_chart,
        'top_skills_chart': top_skills_chart,
        'age_distribution_chart': age_distribution_chart,
    }
    all_skills = Skill.objects.all()
    all_districts = Address.objects.values_list('district', flat=True).distinct()
    context.update({
        'all_skills': all_skills,
        'all_districts': all_districts,
        'filtered_profiles': get_filtered_profiles(request),  # Add this line
    })
    return render(request, 'superuser_dashboard.html', context)

def subscription_plans_view(request):
    return render(request,'subscription_plans.html')

def landing_page_view(request):
    profile1 =Level1Profile.objects.all()[:3]
    profile2 = Level2Profile.objects.all()[:4]
    profile3 = Level3Profile.objects.all()[:4]
    companies = CompanyProfile.objects.all()[:6]
    context ={
        'profile1':profile1,
        'profile2':profile2,
        'profile3':profile3,
        'companies':companies,
    }

    return render(request,'landing_page.html',context)
def allprofiles2_view(request):
    profiles2_list = Level2Profile.objects.all()
    paginator = Paginator(profiles2_list,8)
    page_number = request.GET.get('page')
    profiles2 = paginator.get_page(page_number)

    context = {
        'profiles2':profiles2
    }
    return render(request,'allprofile2.html',context)

# Profile details

@login_required
def level2_profile_detail(request, pk):
    profile = get_object_or_404(Level2Profile, pk=pk)

    if request.method == 'POST' and request.user == profile.user:
        form = Level2ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('level2_profile_detail', pk=pk)
    else:
        form = Level2ProfileForm(instance=profile) if request.user == profile.user else None

    return render(request, 'level2_profile_detail.html', {
        'profile': profile,
        'form': form
    })


#Company profiles
def allcompanyprofiles_view(request):
    companies_list =  CompanyProfile.objects.all()
    paginator = Paginator(companies_list,6)
    page_number = request.GET.get('page')
    companies = paginator.get_page(page_number)

    context = {
        'companies':companies
    }
    return render(request,'allcompanyprofiles.html',context)

def allprofiles3_view(request):
    profiles3_list = Level3Profile.objects.all()
    paginator = Paginator(profiles3_list,8)
    page_number = request.GET.get('page')
    profiles3 = paginator.get_page(page_number)

    context = {
        'profiles3':profiles3
    }
    return render(request,'allprofile3.html',context)

def allprofiles1_view(request):
    query = request.GET.get('q')  # global search
    location = request.GET.get('location')
    skill = request.GET.get('skill')
    category = request.GET.get('category')

    profiles1_list = Level1Profile.objects.all()

    if query:
        profiles1_list = profiles1_list.filter(
            Q(user__username__icontains=query) |
            Q(phone__icontains=query) |
            Q(skills__name__icontains=query) |
            Q(address__province__icontains=query) |
            Q(address__district__icontains=query) |
            Q(address__cell__icontains=query) |
            Q(address__village__icontains=query)
        ).distinct()

    if location:
        profiles1_list = profiles1_list.filter(
            Q(address__province__icontains=location) |
            Q(address__district__icontains=location) |
            Q(address__cell__icontains=location) |
            Q(address__village__icontains=location)
        ).distinct()

    if skill:
        profiles1_list = profiles1_list.filter(skills__name__icontains=skill).distinct()

    if category:
        profiles1_list = profiles1_list.filter(skills__category__name__icontains=category).distinct()

     # Form logic for the logged-in user's profile
    try:
        user_profile = Level1Profile.objects.get(user=request.user)
    except Level1Profile.DoesNotExist:
        user_profile = None

    if request.method == 'POST' and user_profile:
        form = Level1ProfileForm(request.POST, request.FILES, instance=user_profile)
        if form.is_valid():
            form.save()
            return redirect('allprofiles1')  # reload the page after saving
    else:
        form = Level1ProfileForm(instance=user_profile) if user_profile else None

    # Pagination
    paginator = Paginator(profiles1_list, 6)
    page_number = request.GET.get('page')
    profiles1 = paginator.get_page(page_number)

    # Pass filter options to the template
    categories = Category.objects.all()
    skills = Skill.objects.all()

    context = {
        'profiles1': profiles1,
        'query': query,
        'categories': categories,
        'skills': skills,
        'form':form,
    }

    return render(request, 'allprofile1.html', context)


    

    #_____________________Liking profile________________________________
#Profile 1
@require_POST
@login_required
def like_profile(request, profile_id):
    profile = get_object_or_404(Level1Profile, id=profile_id)
    liked = False
    
    if request.user in profile.likes.all():
        profile.likes.remove(request.user)
    else:
        profile.likes.add(request.user)
        liked = True
    
    return JsonResponse({
        'liked': liked,
        'total_likes': profile.likes.count()
    })




#Like_Level_2_Profile
@require_POST
@login_required
def like_level2_profile(request, profile_id):
    profile = get_object_or_404(Level2Profile, id=profile_id)
    liked = False
    
    if request.user in profile.likes.all():
        profile.likes.remove(request.user)
    else:
        profile.likes.add(request.user)
        liked = True

    return JsonResponse({
        'liked': liked,
        'total_likes': profile.likes.count()
    })

#________________________________Sending messages________________________________

@login_required(login_url='login')
def inbox_view(request):
    profile = request.user.profile
    messagesRequests = profile.messages.all()
    unreadCount = messagesRequests.filter(is_read=False).count()
    context = {
        'messagesRequests':messagesRequests ,
        'unreadCount':unreadCount, 

    }

    return render(request, 'inbox.html',context)


@login_required(login_url='login')
def new_inbox_view(request):
    profile = request.user.profile
    messagesRequests = profile.messages.all()
    unreadCount = messagesRequests.filter(is_read=False).count()
    context = {
        'messagesRequests':messagesRequests ,
        'unreadCount':unreadCount, 

    }

    return render(request, 'new_inbox.html',context)

@login_required(login_url='login')
def viewMessage(request, pk):
    profile = request.user.profile
    message = profile.messages.get(pk=pk)
    if not message.is_read:
        message.is_read = True
        message.save()
    context = {
        'message': message
    }
    return render(request, 'message.html', context)


def createMessage(request, pk):
    recipient = Profile.objects.get(id=pk)
    form = MessageForm()

    try:
        sender = request.user.profile
    except:
        sender = None

    if request.method == 'POST':
        form = MessageForm(request.POST)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = sender
            message.recipient = recipient

            if sender:
                message.name = sender.name
                message.email = sender.email
            message.save()

            messages.success(request, 'Your message was successfully sent!')
            return redirect('allprofiles2')

    context = {'recipient': recipient, 'form': form}
    return render(request, 'message_form.html', context)

@staff_member_required
def export_users_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'
    writer = csv.writer(response)
    writer.writerow(['Username', 'Email', 'Phone', 'Address', 'Profile Level', 'Skills', 'Date Joined'])

    profiles = get_filtered_profiles(request)
    for profile in profiles:
        user = getattr(profile, 'user', None)
        username = user.username if user else ''
        email = user.email if user else ''
        phone = getattr(profile, 'phone', '')
        address = ''
        if getattr(profile, 'address', None):
            addr = profile.address
            address = f"{addr.village}, {addr.cell}, {addr.sector}, {addr.district}, {addr.province}"
        profile_level = getattr(profile, 'level', '')
        skills = ''
        if hasattr(profile, 'skills'):
            skills = ', '.join([s.name for s in profile.skills.all()])
        date_joined = user.date_joined.strftime('%Y-%m-%d %H:%M') if user else ''
        writer.writerow([username, email, phone, address, profile_level, skills, date_joined])
    return response

@staff_member_required
def export_users_excel(request):
    import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RwandaTempConnect Users"

    ws.merge_cells('A1:G1')
    ws['A1'] = "RwandaTempConnect Hub - User Report"
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = openpyxl.styles.PatternFill("solid", fgColor="4FACFE")
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A2:G2')
    ws['A2'] = "This document provides a summary of registered users on RwandaTempConnect Hub."
    ws['A2'].font = openpyxl.styles.Font(size=11, italic=True, color="333333")
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    headers = ['Username', 'Email', 'Phone', 'Address', 'Profile Level', 'Skills', 'Date Joined']
    ws.append(headers)
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4FACFE")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    profiles = get_filtered_profiles(request)
    for profile in profiles:
        user = getattr(profile, 'user', None)
        username = user.username if user else ''
        email = user.email if user else ''
        phone = getattr(profile, 'phone', '')
        address = ''
        if getattr(profile, 'address', None):
            addr = profile.address
            address = f"{addr.village}, {addr.cell}, {addr.sector}, {addr.district}, {addr.province}"
        profile_level = getattr(profile, 'level', '')
        skills = ''
        if hasattr(profile, 'skills'):
            skills = ', '.join([s.name for s in profile.skills.all()])
        date_joined = user.date_joined
        if user and is_aware(date_joined):
            date_joined = date_joined.replace(tzinfo=None)
        ws.append([
            username, email, phone, address, profile_level, skills,
            date_joined.strftime('%Y-%m-%d %H:%M') if user else ''
        ])

    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="E3F0FF")
            else:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF")
            cell.alignment = openpyxl.styles.Alignment(vertical="center")

    footer_row = ws.max_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=7)
    ws.cell(row=footer_row, column=1).value = "Generated by RwandaTempConnect Hub Admin Dashboard"
    ws.cell(row=footer_row, column=1).font = openpyxl.styles.Font(italic=True, color="888888")
    ws.cell(row=footer_row, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def export_users_pdf(request):
    import os
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="users_report.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=120, height=60))
        elements.append(Spacer(1, 12))

    title = Paragraph("<b>RwandaTempConnect Hub - User Report</b>", styles['Title'])
    subtitle = Paragraph(
        "This document provides a summary of registered users on RwandaTempConnect Hub. "
        "It includes key user information for administrative and analytical purposes.",
        styles['Normal']
    )
    elements.extend([title, Spacer(1, 8), subtitle, Spacer(1, 20)])

    data = [['Username', 'Email', 'Phone', 'Address', 'Profile Level', 'Skills', 'Date Joined']]

    profiles = get_filtered_profiles(request)
    for profile in profiles:
        user = getattr(profile, 'user', None)
        username = user.username if user else ''
        email = user.email if user else ''
        phone = getattr(profile, 'phone', '')
        address = ''
        if getattr(profile, 'address', None):
            addr = profile.address
            address = f"{addr.village}, {addr.cell}, {addr.sector}, {addr.district}, {addr.province}"
        profile_level = getattr(profile, 'level', '')
        skills = ''
        if hasattr(profile, 'skills'):
            skills = ', '.join([s.name for s in profile.skills.all()])
        date_joined = user.date_joined.strftime('%Y-%m-%d %H:%M') if user else ''
        data.append([username, email, phone, address, profile_level, skills, date_joined])

    table = Table(data, repeatRows=1, colWidths=[60, 100, 70, 120, 60, 100, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4facfe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    footer = Paragraph(
        "<i>Generated by RwandaTempConnect Hub Admin Dashboard</i>",
        styles['Normal']
    )
    elements.append(footer)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required(login_url='login')
def mark_message_read(request, pk):
    if request.method == 'POST':
        profile = request.user.profile
        try:
            message = profile.messages.get(pk=pk)
            if not message.is_read:
                message.is_read = True
                message.save()
            return JsonResponse({'success': True})
        except Message.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Message not found'}, status=404)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

def get_filtered_profiles(request):
    level = request.GET.get('level')
    skill_id = request.GET.get('skill')
    location = request.GET.get('location')
    age_range = request.GET.get('age_range')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Collect all profiles based on level
    profiles = []
    if level == 'level1' or not level:
        profiles += list(Level1Profile.objects.all())
    if level == 'level2' or not level:
        profiles += list(Level2Profile.objects.all())
    if level == 'level3' or not level:
        profiles += list(Level3Profile.objects.all())
    if level == 'company' or not level:
        profiles += list(CompanyProfile.objects.all())

    # Filter by skill
    if skill_id:
        skill = Skill.objects.filter(id=skill_id).first()
        if skill:
            profiles = [p for p in profiles if hasattr(p, 'skills') and skill in p.skills.all()]

    # Filter by location (district)
    if location:
        profiles = [p for p in profiles if p.address and p.address.district == location]

    # Filter by age range (for Level1/2/3 only, assuming user.profile has date_of_birth or user has date_joined)
    if age_range:
        now = datetime.now().date()
        age_bounds = {
            '18-24': (18, 24),
            '25-34': (25, 34),
            '35-44': (35, 44),
            '45+': (45, 150),
        }
        min_age, max_age = age_bounds.get(age_range, (0, 150))
        filtered = []
        for p in profiles:
            user = getattr(p, 'user', None)
            if user and hasattr(user, 'profile') and hasattr(user.profile, 'date_of_birth') and user.profile.date_of_birth:
                dob = user.profile.date_of_birth
                age = (now - dob).days // 365
                if min_age <= age <= max_age:
                    filtered.append(p)
            else:
                # If no DOB, skip or include based on your policy
                pass
        profiles = filtered if filtered else profiles

    # Filter by join date (user.date_joined)
    if start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        profiles = [p for p in profiles if hasattr(p, 'user') and p.user.date_joined.date() >= start.date()]
    if end_date:
        end = datetime.strptime(end_date, '%Y-%m-%d')
        profiles = [p for p in profiles if hasattr(p, 'user') and p.user.date_joined.date() <= end.date()]

    return profiles










